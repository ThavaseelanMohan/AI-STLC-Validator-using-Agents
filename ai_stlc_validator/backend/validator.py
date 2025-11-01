import openpyxl
from openpyxl.styles import PatternFill, Font
import pandas as pd
from similarity import get_similarity_score
from transformers import pipeline
import os

# Load an open model (you can swap this easily)
# For lightweight environments, use: "google/flan-t5-base"
validator_model = pipeline("text2text-generation", model="google/flan-t5-base")

def validate_documents(mapping_df, testcase_df):
    results = []

    for _, row in mapping_df.iterrows():
        req_id = row.get("Requirement_ID", "")
        requirement = row.get("Requirement_Description", "")

        best_tc, best_score = get_similarity_score(requirement, testcase_df)

        # Use open LLM to reason about correctness
        prompt = f"""
Requirement: {requirement}
Test Case: {best_tc}

Determine if the test case correctly satisfies the requirement.
If not, describe the mismatch and suggest a correction.
Respond in one short paragraph.
"""

        llm_output = validator_model(prompt, max_length=256, do_sample=False)[0]["generated_text"]

        results.append({
            "Requirement_ID": req_id,
            "Requirement": requirement,
            "Matched_TestCase": best_tc,
            "Similarity_Score": best_score,
            "AI_Validation": llm_output
        })

    report = pd.DataFrame(results)
    os.makedirs("reports", exist_ok=True)
    report_path = "reports/validation_report.xlsx"
    report.to_excel(report_path, index=False)

    def colorize_excel(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # Header style
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    # Row coloring based on result
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        result = row[4].value
        if "✅" in str(result):
            color = "C6EFCE"  # Green
        elif "⚠️" in str(result):
            color = "FFF2CC"  # Yellow
        else:
            color = "F8CBAD"  # Red
        for cell in row:
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    wb.save(path)
    
    return report_path
